from openpyxl import Workbook,load_workbook
from openpyxl.styles import NamedStyle,Alignment,Font,PatternFill,Border,Side,colors
from openpyxl.worksheet.datavalidation import DataValidation
import weather
import os,time
import pandas as pd
while True:
	#list some city names
	
	cities=["mumbai","hyderabad","bengaluru","delhi","kolkata","chennai","new york","london","bangkok"]
	#load excel
	try:
		df=pd.read_excel("temparatures.xlsx")
		if df["Update"][0].lower()=="no":
			time.sleep(15)
			continue

	except:
		pass
	
	wb=Workbook()

	ws=wb.active

	ws.title="Main Sheet"

	headers=["Sl No","City","Temparature","Update"]

	ws.column_dimensions['B'].width=20
	ws.column_dimensions['C'].width=17

	#create a style
	heading_style=NamedStyle(name="heading_style")
	heading_style.font=Font(color=colors.BLACK,size=14,bold=True)
	heading_style.alignment=Alignment(horizontal="center",vertical="center")
	heading_style.fill=PatternFill("solid",fgColor="FFA500")
	bd = Side( "thin",color="000000")
	heading_style.border=Border(top=bd,right=bd,bottom=bd,left=bd)

	header_index=0
	for row in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=4):

		for cell in row:
			cell.value=headers[header_index]
			cell.style=heading_style
			header_index+=1
	#write serial numbers
	_city_no=1
	for column in ws.iter_cols(min_col=1,max_col=1,min_row=2,max_row=len(cities)+1):
		for cell in column:
			cell.value=_city_no
			_city_no+=1

	#write city names
	_city_no=0
	for column in ws.iter_cols(min_row=2,min_col=2,max_row=1+len(cities),max_col=2):
		for cell in column:
			cell.value=cities[_city_no]
			_city_no+=1

	#note temparature
	temparatures=weather.return_temp_to_excel(cities)
	_city_no=0
	for column in ws.iter_cols(min_row=2,min_col=3,max_row=1+len(cities),max_col=3):
		
		for cell in column:
			cell.value=temparatures[_city_no]
			_city_no+=1

	#create new sheet for city names

	city_sheet=wb.create_sheet()

	wb.active=city_sheet

	city_sheet.title="City Names"

	city_sheet["A1"]="Sl No"
	city_sheet["A1"].style=heading_style
	city_sheet["B1"].style=heading_style

	_city_no=1
	for column in city_sheet.iter_cols(min_col=1,max_col=1,min_row=2,max_row=len(cities)+1):
		for cell in column:
			cell.value=_city_no
			_city_no+=1

	_city_no=0
	city_sheet["B1"]="City"
	#save the city names in city_sheet
	for column in city_sheet.iter_cols(min_row=2,min_col=2,max_row=1+len(cities),max_col=2):
		for cell in column:
			cell.value=cities[_city_no]
			_city_no+=1

	wb.active=ws

	try:
		wb.save("temparatures.xlsx")
	except:
		pass
	time.sleep(15)